#!/usr/bin/env python
# coding=utf-8
import fnmatch
import os
import shutil
import zipfile
import openpyxl
import numpy as np
from openpyxl.styles import *
from openpyxl.utils import get_column_letter, column_index_from_string

sheet_del_col_num_list = []

# 处理zipfile解压中文乱码
def support_gbk(zip_file: zipfile.ZipFile):
    name_to_info = zip_file.NameToInfo
    # copy map first
    for name, info in name_to_info.copy().items():
        real_name = name.encode('cp437').decode('gbk')
        if real_name != name:
            info.filename = real_name
            del name_to_info[name]
            name_to_info[real_name] = info
    return zip_file

# 扫描当前目录，解压zip包，返回解压后目录
def scan_handle_dircetory():
    # 保存所有需要处理的目录
    directory_list = []
    # 获取当前目录
    current_directory = os.getcwd()
    # 遍历当前目录下的所有压缩包，并解压
    for file_name in os.listdir(current_directory):
        # 检查文件是否是zip压缩包
        if file_name.endswith('.zip'):
            # 构建zip文件的绝对路径
            zip_file_path = os.path.join(current_directory, file_name)

            # 打开zip文件
            with support_gbk(zipfile.ZipFile(zip_file_path, 'r')) as zip_ref:
                # 解压缩当前压缩包到当前目录
                zip_ref.extractall(current_directory)
            directory_list.append(file_name)
    return directory_list

# 遍历解压目录下的所有文件，返回所有小题分excel文件
def handle_excel_list():
    excel_list = []
    for item in scan_handle_dircetory():
        target_directory = item[0:-4]
        # 遍历目录下的所有文件
        for root, dirs, files in os.walk(target_directory):
            for file_name in files:
                # 判断是否为 Excel 文件，且文件名中包含 "各科小题分" 字段
                if fnmatch.fnmatch(file_name, '*各科小题分*.xlsx'):
                    file_path = os.path.join(root, file_name)
                    excel_list.append(file_path)
    return excel_list

# 判断是否存在 del.txt 文件
def if_with_exist():
    # 获取当前目录
    current_directory = os.getcwd()

    # 构建handle.txt的路径
    handle_file_path = os.path.join(current_directory, 'del.txt')

    # 判断handle.txt是否存在
    if os.path.isfile(handle_file_path):
        return handle_file_path
    else:
        return None

# 读取 del.txt 文件，获取需要删除的列
def del_cols_list(path):
    # 保存需要删除的列
    del_cols = {}

    f = open(path, encoding="utf-8")
    for line in f.readlines():
        if len(line) != 1:
            line = line.strip('\n')
            notes = line.split(" ")
            t = []
            for i, j in enumerate(notes):
                if i != 0:
                    t.append(j)
                    del_cols.update({str(notes[0]): t})
    return del_cols

# 取消已经合并的单元格
def cancel_all_merged(sheet):
    # 合并单元格的位置信息，可迭代对象（单个是一个 'openpyxl.worksheet.cell_range.CellRange'对象）print后就是excel坐标信息
    m_list = sheet.merged_cells
    cr = []
    rr = []
    for m_area in m_list:
        # 合并单元格的起始行坐标，终止行坐标
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # 列合并信息提取
        if c2 - c1 > 0:
            cr.append((r1, r2, c1, c2))
        # 行合并信息提取
        if r2 - r1 > 0:
            rr.append((r1, r2, c1, c2))
    # 取消所有列的合并
    for c in cr:
        sheet.unmerge_cells(start_row=c[0], end_row=c[1], start_column=c[2], end_column=c[3])
    # 取消所有行的合并
    for r in rr:
        sheet.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])
    return sheet

# 收集需要合并的单元格范围
def collect_merged_cells(sheet):
    # sheet中需要合并的范围（第一行，针对列）
    merge_col_area = []
    # 合并后的列内容预存
    merge_col_area_name = []
    # 初始化 merge_col_area 和 merge_col_area_name
    for i in range(1, sheet.max_column):
        value1 = sheet.cell(row=1, column=i).value
        value2 = sheet.cell(row=1, column=i + 1).value
        if value1 != None and value2 == None:
            merge_col_area.append(i)
            merge_col_area_name.append(value1)
    merge_col_area.append(sheet.max_column)
    return merge_col_area, merge_col_area_name

# 删除列
def do_cols_delete(sheet, obj_del_col, merge_col_area):
    # 每一个sheet中删除了多少列
    counts = 0
    # 对 sheet 进行列处理(删除)
    for sheet_del_col_idx in range(len(obj_del_col)):
        # 处理不需要列合并的前几个列（是否需要删除列）
        if obj_del_col[sheet_del_col_idx] < merge_col_area[0]:
            sheet.delete_cols(obj_del_col[sheet_del_col_idx])
            counts += 1
            # 删除列后，所有的后续列号会变化，需要调整
            merge_col_area = (np.array(merge_col_area) - 1).tolist()
            obj_del_col = (np.array(obj_del_col) - 1).tolist()
        else:
            for t in range(len(merge_col_area) - 1):
                if merge_col_area[t] <= obj_del_col[sheet_del_col_idx] < merge_col_area[t + 1]:
                    sheet.delete_cols(obj_del_col[sheet_del_col_idx])
                    counts += 1
                    obj_del_col = (np.array(obj_del_col) - 1).tolist()
                    for tt in range(t + 1, len(merge_col_area)):
                        merge_col_area[tt] = merge_col_area[tt] - 1
    sheet_del_col_num_list = counts
    return sheet, sheet_del_col_num_list, merge_col_area

# 根据 merge_col_area 给出的合并范围，合并单元格
def do_cols_merge(sheet, merge_col_area, merge_col_area_name):
    # 取出合并的内容并赋值
    for cell_idx, cell_val in enumerate(merge_col_area):
        if cell_idx < len(merge_col_area_name):
            sheet.cell(row=1, column=cell_val, value=merge_col_area_name[cell_idx])

    # 列合并
    for merge_col_area_idx in range(len(merge_col_area) - 1):
        value1 = int(merge_col_area[merge_col_area_idx])
        if merge_col_area_idx == len(merge_col_area) - 2:
            value2 = int(merge_col_area[merge_col_area_idx + 1])
        else:
            value2 = int(merge_col_area[merge_col_area_idx + 1] - 1)
        sheet.merge_cells(start_row=1, end_row=1, start_column=value1, end_column=value2)
    sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
    sheet.merge_cells(start_row=1, end_row=2, start_column=2, end_column=2)
    return sheet

# 删除给定的列
def delete_cols(sheet, del_cols):
    sheet = cancel_all_merged(sheet)
    merge_col_area, merge_col_area_name = collect_merged_cells(sheet)

    # 创建 title_dict 字典
    title_dict = {key: [int(column_index_from_string(col)) if isinstance(col, str) else col for col in cols]
                  for key, cols in del_cols.items()}
    # 检查 sheet.title 是否在 title_dict 的键中，获取对应的处理结果
    # 通过 title_dict 将需要删除的 ['A', 'C', 'E'] 转为 [1, 3, 5]
    obj_del_col = title_dict.get(sheet.title, [])
    # 对需要删除的学科的列进行排序
    obj_del_col = sorted(obj_del_col)

    # 删除相关的列
    sheet, sheet_del_col_num_list, merge_col_area= do_cols_delete(sheet, obj_del_col, merge_col_area)
    # 删除之后，对剩余部分进行合并
    sheet =  do_cols_merge(sheet, merge_col_area, merge_col_area_name)
    return sheet, sheet_del_col_num_list

# 打印配置
def print_config(sheet, sheet_del_col_num_list):
    # 获取行列个数
    maxcol = sheet.max_column
    maxrow = sheet.max_row
    # 设置打印区域
    maxcol = sheet.max_column
    maxrow = sheet.max_row
    area = "A1" + ":" + get_column_letter(maxcol - sheet_del_col_num_list) + str(maxrow)
    sheet.print_area = area
    # 设置打印A3横向
    sheet.set_printer_settings(sheet.PAPERSIZE_A3, sheet.ORIENTATION_LANDSCAPE)
    # 设置打印样式
    font_12 = Font(
        name='Arial',
        size=9,
        bold=True
    )
    font_ot = Font(name='Arial', size=9)
    for i in range(1, 3):
        for j in range(1, maxcol + 1):
            sheet.cell(row=i, column=j).font = font_12
            sheet.cell(row=i, column=j).fill = PatternFill("solid", fgColor="DBDBDB")
    for i in range(3, maxrow + 1):
        for j in range(1, maxcol + 1):
            sheet.cell(row=i + 1, column=j + 1).font = font_ot
    # 设置所有单元格居中
    for i in range(1, maxrow + 1):
        for j in range(1, maxcol + 1):
            sheet.cell(row=i, column=j).alignment = Alignment(horizontal='center', vertical='center')
    # 页面居中
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # 设置页边距
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.2
    sheet.page_margins.bottom = 0.2
    sheet.page_margins.header = 0.2
    sheet.page_margins.footer = 0.2
    # 设置页眉高度
    sheet.oddHeader.left.size = 0
    sheet.oddHeader.center.size = 0
    sheet.oddHeader.right.size = 0
    # 设置页脚高度
    sheet.oddFooter.left.size = 0
    sheet.oddFooter.center.size = 0
    sheet.oddFooter.right.size = 0
    # 缩放页面，缩放到一页
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToHeight = True
    sheet.page_setup.fitToWidth = True
    return sheet

def adjust_width_high(sheet):
    # A4 横向 长 * 宽 297 * 210 mm
    # 行 1mm = 2.83 磅
    # 列 1mm = 4.18 字符
    row = 841 / sheet.max_row
    column = 230 / sheet.max_column
    for item in range(1, sheet.max_row):
        sheet.row_dimensions[item].height = row
    for item in range(1, sheet.max_column):
        sheet.column_dimensions[get_column_letter(item)].width = column
    return sheet

# 保存excel
def save_excel(workbook, old_path, new_path):
    # 提取上一级目录路径
    directory_path = "\\".join(new_path.split("\\")[:-1])
    # 如果目录不存在则创建目录
    os.makedirs(directory_path, exist_ok=True)
    workbook.save(new_path)

if __name__ == "__main__":
    handle_txt = if_with_exist()
    directory_list = scan_handle_dircetory()

    if handle_txt == None:
        print("del.txt 文件不存在，请创建后在执行！")
        input("请输入任意字符回车结束：")
        exit()
    if len(directory_list) == 0:
        print("不存在小题分压缩包 *.zip ，请将压缩包拷贝到 小题分神器.ext 同级目录下！")
        input("请输入任意字符回车结束：")
        exit()
    target_dir = "小题分"
    while os.path.exists(target_dir):
        if os.path.isdir(target_dir):
            target_dir += "x"
        else:
            break

    old_excels = handle_excel_list()
    new_excels = [(target_dir + "\\") + s for s in old_excels]
    del_cols = del_cols_list(handle_txt)

    for index, path in enumerate(old_excels):
        workbook = openpyxl.load_workbook(path)
        sheet_names = workbook.sheetnames
        for item, sheet_name in enumerate(sheet_names):
            # 获取工作表对象
            sheet = workbook[sheet_name]
            # 删除列
            sheet, sheet_del_col_num_list = delete_cols(sheet, del_cols)
            sheet = adjust_width_high(sheet)
            # 设置打印配置
            sheet = print_config(sheet, sheet_del_col_num_list)
        print("processing excel：", path);
        # 保存所有excel
        save_excel(workbook, path, new_excels[index])

    # 删除解压后的文件
    dir_list_without_zip = [item.rstrip('.zip') for item in directory_list]
    for index_dir in dir_list_without_zip:
        shutil.rmtree(index_dir)
    processe_end = input("请输入任意字符回车结束：")

